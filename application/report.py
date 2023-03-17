from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from openpyxl.worksheet.dimensions import ColumnDimension
import string


def create_report(date_insert, res_list):
    alphabit = list(string.ascii_uppercase)
    dest_filename = 'application/static/report/Отчет на ' + str(date_insert) + '.xlsx'
    wb = Workbook()
    # FIRST SHEET
    ws = wb.active
    ws.title = 'Личный состав'
    # Styles
    ft = Font(bold=True, size=14)
    ft_2 = Font(bold=True, size=10)
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thick = Side(border_style="thin")
    double = Side(border_style="double")
    br = Border(top=double, left=double, right=double, bottom=double)
    br2 = Border(top=thick, left=thick, right=thick, bottom=thick)
    for col_letter in alphabit[:15]:
        ws.column_dimensions[col_letter] = ColumnDimension(ws, bestFit=True, customWidth=True)
        if col_letter in ['A', 'J']:
            ws.column_dimensions[col_letter].width = 19
        else:
            ws.column_dimensions[col_letter].width = 15
        ws[col_letter + '1'].border = br
        ws[col_letter + '2'].border = br
        ws[col_letter + '1'].font = ft
        ws[col_letter + '2'].font = ft
        ws[col_letter + '1'].alignment = al
        ws[col_letter + '2'].alignment = al

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45

    # Merge Cells OF HEADER
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:J1')
    ws.merge_cells('K1:K2')
    ws.merge_cells('L1:O1')

    # Assignment
    ws['A1'] = 'Наименование пожарных подразделений '
    ws['B1'] = 'Штат ПЧ (СПЧ,ПП)'
    ws['C1'] = 'В карауле по списку л/с'
    ws['D1'] = 'Налицо личного состава'
    ws['D2'] = 'Всего'
    ws['E2'] = 'Расчет'
    ws['F2'] = 'Начальники караулов'
    ws['G2'] = 'Командиры отделений'
    ws['H2'] = 'Водители'
    ws['I2'] = 'Пожарные'
    ws['J2'] = 'Диспетчеров (радиотелефонистов)'
    ws['K1'] = 'Газодымозащитники/ аппараты'
    ws['L1'] = 'Отсутствуют'
    ws['L2'] = 'Отпуск учебный/декрет.'
    ws['M2'] = 'Больные'
    ws['N2'] = 'Командировка'
    ws['O2'] = 'Другие причины'

    row = 3
    for r in res_list:
        ws.cell(column=1, row=row, value=r['dep_unique_id']).border = br2
        ws.cell(column=2, row=row, value=r['num_fighter']).border = br2
        ws.cell(column=3, row=row, value=r['karaul_list']).border = br2
        ws.cell(column=4, row=row, value=r['karaul_all']).border = br2
        ws.cell(column=5, row=row, value=r['karaul_fighter']).border = br2
        ws.cell(column=6, row=row, value=r['head']).border = br2
        ws.cell(column=7, row=row, value=r['commander']).border = br2
        ws.cell(column=8, row=row, value=r['drivers']).border = br2
        ws.cell(column=9, row=row, value=r['fire_fighter']).border = br2
        ws.cell(column=10, row=row, value=r['call_assistant']).border = br2
        ws.cell(column=11, row=row, value=r['gdzs']).border = br2
        ws.cell(column=12, row=row, value=r['vacation']).border = br2
        ws.cell(column=13, row=row, value=r['sick']).border = br2
        ws.cell(column=14, row=row, value=r['business_trip']).border = br2
        ws.cell(column=15, row=row, value=r['others']).border = br2
        row += 1

    # SECOND SHEET

    ws1 = wb.create_sheet("Пожарная техника")

    for col_letter in alphabit[:30]:
        ws1.column_dimensions[col_letter] = ColumnDimension(ws1, bestFit=True, customWidth=True)
        ws1.column_dimensions[col_letter].width = 19
        ws1[col_letter + '1'].border = br
        ws1[col_letter + '2'].border = br
        ws1[col_letter + '3'].border = br
        ws1[col_letter + '1'].font = ft
        ws1[col_letter + '2'].font = ft
        ws1[col_letter + '3'].font = ft_2
        ws1[col_letter + '1'].alignment = al
        ws1[col_letter + '2'].alignment = al
        ws1[col_letter + '3'].alignment = al
    alphabit_next = ['aa', 'ab']
    for col_letter in alphabit_next:
        ws1.column_dimensions[col_letter] = ColumnDimension(ws1, bestFit=True, customWidth=True)
        ws1.column_dimensions[col_letter].width = 19
        ws1[col_letter + '1'].border = br
        ws1[col_letter + '2'].border = br
        ws1[col_letter + '3'].border = br
        ws1[col_letter + '1'].font = ft
        ws1[col_letter + '2'].font = ft
        ws1[col_letter + '3'].font = ft_2
        ws1[col_letter + '1'].alignment = al
        ws1[col_letter + '2'].alignment = al
        ws1[col_letter + '3'].alignment = al

    ws1.row_dimensions[1].height = 20
    ws1.row_dimensions[2].height = 65

    # Merge Cells OF HEADER

    # Assignment
    ws1['A1'] = 'Наименование пожарных подразделений'
    ws1['B1'] = 'ПОЖАРНАЯ ТЕХНИКА'

    ws1['B2'] = 'В расчете'
    ws1['F2'] = 'В резерве'
    ws1['J2'] = 'На ремонте'

    ws1.merge_cells(f'B{1}:K{1}')

    ws1.merge_cells(f'B{2}:E{2}')
    ws1.merge_cells(f'F{2}:I{2}')
    ws1.merge_cells(f'J{2}:K{2}')
    ws1['b2'].fill = PatternFill(patternType='solid', fgColor='4EADEA')
    ws1['f2'].fill = PatternFill(patternType='solid', fgColor='FFFF54')
    ws1['j2'].fill = PatternFill(patternType='solid', fgColor='DF9A9E')
    ws1.merge_cells(f'A{1}:A{3}')

    ws1.merge_cells('L1:L3')
    ws1.merge_cells('M1:M3')

    ws1.merge_cells('N1:O2')

    ws1.merge_cells('P1:R2')

    ws1.merge_cells('S1:AB1')
    ws1.merge_cells('S2:T2')

    ws1.merge_cells('U2:U3')
    ws1.merge_cells('V2:V3')
    ws1.merge_cells('W2:W3')
    ws1.merge_cells('X2:X3')
    ws1.merge_cells('Y2:Y3')
    ws1.merge_cells('Z2:Z3')
    ws1.merge_cells('AA2:AA3')
    ws1.merge_cells('AB2:AB3')

    ws1['L1'] = 'пена на складе'
    ws1['M1'] = 'гидрокостюм'

    ws1['N1'] = 'Мотопомбы'
    ws1['N3'] = 'в боевом расчете'
    ws1['O3'] = 'на ремонте'

    ws1['P1'] = 'Рукова'
    ws1['P3'] = '77мм'
    ws1['Q3'] = '66мм'
    ws1['R3'] = '51мм'

    ws1['S1'] = 'Вооружение в расчетах'
    ws1['S2'] = 'Лафетные сволы'
    ws1['S3'] = 'стационарные'
    ws1['T3'] = 'переносные'

    ws1['U2'] = 'ГПС-600'
    ws1['V2'] = '«Пурга»'
    ws1['W2'] = 'переносные радиостанции'
    ws1['X2'] = 'электрофонарь'
    ws1['Y2'] = 'прожектор'
    ws1['Z2'] = 'ТОК'
    ws1['AA2'] = 'Л-1'
    ws1['AB2'] = 'спасательные веревки'

    ws1['B3'] = 'Тип основного пожарного автомобиля'
    ws1['B3'].fill = PatternFill(patternType='solid', fgColor='4EADEA')
    ws1['C3'] = 'Марка специального пожарного автомобиля'
    ws1['C3'].fill = PatternFill(patternType='solid', fgColor='4EADEA')
    ws1['D3'] = 'ГСМ'
    ws1['D3'].fill = PatternFill(patternType='solid', fgColor='4EADEA')
    ws1['E3'] = 'Пена'
    ws1['E3'].fill = PatternFill(patternType='solid', fgColor='4EADEA')

    ws1['F3'] = 'Тип основного пожарного автомобиля'
    ws1['F3'].fill = PatternFill(patternType='solid', fgColor='FFFF54')
    ws1['G3'] = 'Марка специального пожарного автомобиля'
    ws1['G3'].fill = PatternFill(patternType='solid', fgColor='FFFF54')
    ws1['H3'] = 'ГСМ'
    ws1['H3'].fill = PatternFill(patternType='solid', fgColor='FFFF54')
    ws1['I3'] = 'Пена'
    ws1['I3'].fill = PatternFill(patternType='solid', fgColor='FFFF54')

    ws1['J3'] = 'Тип основного пожарного автомобиля'
    ws1['J3'].fill = PatternFill(patternType='solid', fgColor='DF9A9E')
    ws1['K3'] = 'Марка специального пожарного автомобиля'
    ws1['K3'].fill = PatternFill(patternType='solid', fgColor='DF9A9E')


    row = 4
    for r in res_list:
        tab_row = r['tab_row']
        if tab_row > -1:
            ws1.cell(column=1, row=row, value=r['dep_unique_id']).border = br2
            try:
                ws1.cell(column=12, row=row, value=str(r['foam_stock'][0])).border = br2
            except IndexError:
                ws1.cell(column=12, row=row, value='').border = br2
            try:
                ws1.cell(column=13, row=row, value=str(r['hydra_costume'][0])).border = br2
            except IndexError:
                ws1.cell(column=13, row=row, value='').border = br2
            try:
                ws1.cell(column=14, row=row, value=str(r['motor_pumps_in_combat'][0])).border = br2
            except IndexError:
                ws1.cell(column=14, row=row, value='').border = br2
            try:
                ws1.cell(column=15, row=row, value=str(r['motor_pumps_on_repair'][0])).border = br2
            except IndexError:
                ws1.cell(column=15, row=row, value='').border = br2


            inner_row = row
            for i in range(tab_row + 1):
                try:
                    ws1.cell(column=2, row=inner_row, value=r['counting_brand'][i]).border = br2
                    ws1.cell(column=3, row=inner_row, value=r['counting_model'][i]).border = br2
                    ws1.cell(column=4, row=inner_row, value=r['counting_gsm'][i]).border = br2
                    ws1.cell(column=5, row=inner_row, value=r['counting_foam'][i]).border = br2

                    ws1.cell(column=16, row=inner_row, value=r['counting_sleeves_77'][i]).border = br2
                    ws1.cell(column=17, row=inner_row, value=r['counting_sleeves_66'][i]).border = br2
                    ws1.cell(column=18, row=inner_row, value=r['counting_sleeves_51'][i]).border = br2

                    ws1.cell(column=19, row=inner_row, value=r['counting_fire_monitors_stationary'][i]).border = br2
                    ws1.cell(column=20, row=inner_row, value=r['counting_fire_monitors_portable'][i]).border = br2
                    ws1.cell(column=21, row=inner_row, value=r['counting_gps_600'][i]).border = br2
                    ws1.cell(column=22, row=inner_row, value=r['counting_blizzard'][i]).border = br2
                    ws1.cell(column=23, row=inner_row, value=r['counting_portable_radios'][i]).border = br2
                    ws1.cell(column=24, row=inner_row, value=r['counting_flashlight'][i]).border = br2
                    ws1.cell(column=25, row=inner_row, value=r['counting_spotlight'][i]).border = br2
                    ws1.cell(column=26, row=inner_row, value=r['counting_current'][i]).border = br2
                    ws1.cell(column=27, row=inner_row, value=r['counting_l1'][i]).border = br2
                    ws1.cell(column=28, row=inner_row, value=r['counting_rescue_ropes'][i]).border = br2

                    inner_row += 1
                except IndexError:
                    ws1.cell(column=2, row=inner_row, value='').border = br2
                    ws1.cell(column=3, row=inner_row, value='').border = br2
                    ws1.cell(column=4, row=inner_row, value='').border = br2
                    ws1.cell(column=5, row=inner_row, value='').border = br2
                    inner_row += 1
            inner_row = row
            for i in range(tab_row + 1):
                try:
                    ws1.cell(column=6, row=inner_row, value=r['reserve_brand'][i]).border = br2
                    ws1.cell(column=7, row=inner_row, value=r['reserve_model'][i]).border = br2
                    ws1.cell(column=8, row=inner_row, value=r['reserve_gsm'][i]).border = br2
                    ws1.cell(column=9, row=inner_row, value=r['reserve_foam'][i]).border = br2
                    inner_row += 1
                except IndexError:
                    ws1.cell(column=6, row=inner_row, value='').border = br2
                    ws1.cell(column=7, row=inner_row, value='').border = br2
                    ws1.cell(column=8, row=inner_row, value='').border = br2
                    ws1.cell(column=9, row=inner_row, value='').border = br2
                    inner_row += 1
            inner_row = row
            for i in range(tab_row + 1):
                try:
                    ws1.cell(column=10, row=inner_row, value=r['renovation_brand'][i]).border = br2
                    ws1.cell(column=11, row=inner_row, value=r['renovation_model'][i]).border = br2
                    inner_row += 1
                except IndexError:
                    ws1.cell(column=10, row=inner_row, value='').border = br2
                    ws1.cell(column=11, row=inner_row, value='').border = br2
                    inner_row += 1
            row = row + tab_row + 1
        else:
            ws1.cell(column=1, row=row, value=r['dep_unique_id']).border = br2
            for i in range(2, 8):
                ws1.cell(column=i, row=row, value='').border = br2

            row += 1
    empty_departs = ['ПЧ 5',
                     'ПЧ 6',
                     'ПЧ 7',
                     'ПЧ 8',
                     'ПЧ 9',
                     'ПЧ 10',
                     'ПЧ 11',
                     'ПЧ 12',
                     'ПЧ 13',
                     'ПЧ 14',
                     'ПЧ 15',
                     'ПЧ 16',
                     'ПЧ 17',
                     'ПЧ 18',
                     'ПЧ 19',
                     'Тех часть',
                     'ОСО 10 мкр.',
                     'Зербулак',
                     'Каратау',
                     'Жанатас']
    row += 1
    for depart in empty_departs:
        ws1.cell(column=1, row=row, value=depart).border = br2
        for col in range(2, 29):
            ws1.cell(column=col, row=row, value='').border = br2
        row += 1

    wb.save(filename=dest_filename)


def create_report_emergency(cur_emg, event_journal):
    dest_filename = 'application/static/report/Отчет на ' + str(cur_emg.date_of_call_start.date()) + '.xlsx'
    alphabit = list(string.ascii_uppercase)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Отчет'
    # Styles
    ft = Font(bold=True, size=14)
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al2 = Alignment(vertical="center", wrap_text=True)
    thick = Side(border_style="thin")
    double = Side(border_style="double")
    br = Border(top=double, left=double, right=double, bottom=double)
    br2 = Border(top=thick, left=thick, right=thick, bottom=thick)
    # SET styles
    ws.column_dimensions['A'] = ColumnDimension(ws, bestFit=True, customWidth=True)
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'] = ColumnDimension(ws, bestFit=True, customWidth=True)
    ws.column_dimensions['B'].width = 35
    for i in range(13):
        ws.row_dimensions[i + 1].height = 30
        ws.cell(column=1, row=i + 1).border = br
        ws.cell(column=1, row=i + 1).font = ft
        ws.cell(column=1, row=i + 1).alignment = al
        ws.cell(column=2, row=i + 1).border = br2
        ws.cell(column=2, row=i + 1).alignment = al2

    # Assignment
    ws['A1'] = 'Время начала звонка'
    ws['A2'] = 'Время конца звонка'
    ws['A3'] = 'Номер тел. заявителя'
    ws['A4'] = 'ФИО заявителя'
    ws['A5'] = 'Адрес'
    ws['A6'] = 'Дом/Кв'
    ws['A7'] = 'Категория пожара'
    ws['A8'] = 'Тип выезда'
    ws['A9'] = 'Ранг выезда'
    ws['A10'] = 'Владелец объекта'
    ws['A11'] = 'Пожарная часть'
    ws['A12'] = 'Первичная информация'
    ws['A13'] = 'Руководитель тушение пожара'
    ws['A14'] = 'Превлеченные сили и средства'

    ws.cell(column=2, row=1, value=str(cur_emg.date_of_call_start))
    ws.cell(column=2, row=2, value=str(cur_emg.date_of_call_start))
    ws.cell(column=2, row=3, value=cur_emg.income_call_number)
    ws.cell(column=2, row=4, value=cur_emg.income_call_name)
    ws.cell(column=2, row=5, value=cur_emg.address)
    ws.cell(column=2, row=6, value=cur_emg.house)
    #cur_emg.fire_category.name
    ws.cell(column=2, row=7, value='Пожар')
    #cur_emg.emergency_type.name
    ws.cell(column=2, row=8, value='Тип происшествия')
    #cur_emg.emergency_rank.rank
    ws.cell(column=2, row=9, value='Ранг происшествия')
    #cur_emg.object_owner
    ws.cell(column=2, row=10, value='Владелец объекта')
    col = 2
    for dep in cur_emg.department.all():
        ws.cell(column=col, row=11, value=dep.name).border = br2
        ws.cell(column=col, row=11).alignment = al2
        col += 1
    #'cur_emg.first_info_burning'
    ws.cell(column=2, row=12, value='Первичная информация')
    #cur_emg.staff
    ws.cell(column=2, row=13, value='стафф').border = br2
    row = 14
    for trans in cur_emg.transport.all():
        ws.cell(column=2, row=row, value=trans.brand.brand).border = br2
        ws.cell(column=1, row=row).border = br2
        ws.cell(column=2, row=13).alignment = al2
        row += 1

    # SECOND SHEET

    ws1 = wb.create_sheet("Журнал событий")
    ws1.row_dimensions[1].height = 25

    # SET styles
    ws1.column_dimensions['A'] = ColumnDimension(ws1, bestFit=True, customWidth=True)
    ws1.column_dimensions['B'] = ColumnDimension(ws1, bestFit=True, customWidth=True)
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 75
    ws1.cell(column=1, row=1).border = br
    ws1.cell(column=1, row=1).font = ft
    ws1.cell(column=1, row=1).alignment = al
    ws1.cell(column=2, row=1).border = br
    ws1.cell(column=2, row=1).font = ft
    ws1.cell(column=2, row=1).alignment = al

    # Assignment
    ws1['A1'] = 'Время'
    ws1['B1'] = 'Событие'
    row = 2
    for event in event_journal:
        ws1.cell(column=1, row=row, value=event.date_insert).border = br2
        ws1.cell(column=2, row=row, value=event.event).border = br2
        row += 1

    wb.save(filename=dest_filename)
    return dest_filename
