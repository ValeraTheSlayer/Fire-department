import os
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension


def create_report_emergency(cur_emg, event_journal):
    dest_filename = 'report/Отчет на ' + str(cur_emg.date_of_call_start.date()) + '.xlsx'
    full_path = os.path.join(settings.MEDIA_ROOT, dest_filename)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Отчет'
    # Styles
    ft = Font(bold=True, size=14)
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al2 = Alignment(vertical="center", wrap_text=True)
    thick = Side(border_style="thin")
    double = Side(border_style="double")
    br = Border(top=double, left=double, right=double, bottom=double)
    br2 = Border(top=thick, left=thick, right=thick, bottom=thick)
    # SET styles
    ws.column_dimensions['A'] = ColumnDimension(ws, bestFit=True, customWidth=True)
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'] = ColumnDimension(ws, bestFit=True, customWidth=True)
    ws.column_dimensions['B'].width = 35
    for i in range(13):
        ws.row_dimensions[i + 1].height = 30
        ws.cell(column=1, row=i + 1).border = br
        ws.cell(column=1, row=i + 1).font = ft
        ws.cell(column=1, row=i + 1).alignment = al
        ws.cell(column=2, row=i + 1).border = br2
        ws.cell(column=2, row=i + 1).alignment = al2

    # Assignment
    ws['A1'] = 'Время начала звонка'
    ws['A2'] = 'Время конца звонка'
    ws['A3'] = 'Номер тел. заявителя'
    ws['A4'] = 'ФИО заявителя'
    ws['A5'] = 'Адрес'
    ws['A6'] = 'Категория пожара'
    ws['A7'] = 'Тип выезда'
    ws['A8'] = 'Ранг выезда'
    ws['A9'] = 'Владелец объекта'
    ws['A10'] = 'Пожарная часть'
    ws['A11'] = 'Первичная информация'
    ws['A12'] = 'Руководитель тушения пожара'
    ws['A13'] = 'Привлеченные силы и средства'
    ws.cell(column=2, row=1, value=str(timezone.localtime(cur_emg.date_of_call_start))[:-6])
    ws.cell(column=2, row=2, value=str(timezone.localtime(cur_emg.date_of_call_start))[:-6])
    ws.cell(column=2, row=3, value=cur_emg.income_call_number)
    ws.cell(column=2, row=4, value=cur_emg.income_call_name)
    ws.cell(column=2, row=5, value=cur_emg.address)
    ws.cell(column=2, row=6, value=cur_emg.object_category.name)
    ws.cell(column=2, row=7, value=cur_emg.emergency_type.name)
    ws.cell(column=2, row=8, value=cur_emg.emergency_rank.rank)
    ws.cell(column=2, row=9, value=cur_emg.object_owner)
    col = 2
    for dep in cur_emg.department.all():
        ws.cell(column=col, row=10, value=dep.name).border = br2
        ws.cell(column=col, row=10).alignment = al2
        col += 1
    ws.cell(column=2, row=11, value=cur_emg.first_info_burning)
    ws.cell(column=2, row=12, value=cur_emg.staff.full_name).border = br2
    row = 13
    for trans in cur_emg.transport.all():
        ws.cell(column=2, row=row, value=trans.brand.brand).border = br2
        ws.cell(column=1, row=row).border = br2
        ws.cell(column=2, row=12).alignment = al2
        row += 1

    # SECOND SHEET

    ws1 = wb.create_sheet("Журнал событий")
    ws1.row_dimensions[1].height = 25

    # SET styles
    ws1.column_dimensions['A'] = ColumnDimension(ws1, bestFit=True, customWidth=True)
    ws1.column_dimensions['B'] = ColumnDimension(ws1, bestFit=True, customWidth=True)
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 75
    ws1.cell(column=1, row=1).border = br
    ws1.cell(column=1, row=1).font = ft
    ws1.cell(column=1, row=1).alignment = al
    ws1.cell(column=2, row=1).border = br
    ws1.cell(column=2, row=1).font = ft
    ws1.cell(column=2, row=1).alignment = al

    # Assignment
    ws1['A1'] = 'Время'
    ws1['B1'] = 'Событие'
    row = 2
    for event in event_journal[::-1]:
        ws1.cell(column=1, row=row, value=str(timezone.localtime(event.date_insert))[:-6]).border = br2
        ws1.cell(column=2, row=row, value=event.event).border = br2
        row += 1

    wb.save(filename=full_path)
    return full_path
