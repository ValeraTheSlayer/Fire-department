import os
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension
import string


def create_report_line_note(date_insert, res_list):
    alphabit = list(string.ascii_uppercase)
    dest_filename = 'report/Отчет на ' + str(date_insert) + '.xlsx'
    full_path = os.path.join(settings.MEDIA_ROOT, dest_filename)
    wb = Workbook()
    # FIRST SHEET
    ws = wb.active
    ws.title = 'Личный состав'
    # Styles
    ft = Font(bold=True, size=14)
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thick = Side(border_style="thin")
    double = Side(border_style="double")
    br = Border(top=double, left=double, right=double, bottom=double)
    br2 = Border(top=thick, left=thick, right=thick, bottom=thick)
    for col_letter in alphabit[:15]:
        ws.column_dimensions[col_letter] = ColumnDimension(ws, bestFit=True, customWidth=True)
        if col_letter in ['A', 'J']:
            ws.column_dimensions[col_letter].width = 19
        else:
            ws.column_dimensions[col_letter].width = 15
        ws[col_letter + '1'].border = br
        ws[col_letter + '2'].border = br
        ws[col_letter + '1'].font = ft
        ws[col_letter + '2'].font = ft
        ws[col_letter + '1'].alignment = al
        ws[col_letter + '2'].alignment = al

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45

    # Merge Cells OF HEADER
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:J1')
    ws.merge_cells('K1:K2')
    ws.merge_cells('L1:O1')

    # Assignment
    ws['A1'] = 'Наименование пожарных подразделений'
    ws['B1'] = 'Штат ПЧ (СПЧ,ПП)'
    ws['C1'] = 'В карауле по списку л/с'
    ws['D1'] = 'Налицо личного состава'
    ws['D2'] = 'Всего'
    ws['E2'] = 'Расчет'
    ws['F2'] = 'Начальники караулов'
    ws['G2'] = 'Командиры отделений'
    ws['H2'] = 'Водители'
    ws['I2'] = 'Пожарные'
    ws['J2'] = 'Диспетчеров (радиотелефонистов)'
    ws['K1'] = 'Газодымозащитники/ аппараты'
    ws['L1'] = 'Отсутствуют'
    ws['L2'] = 'Отпуск учебный/декрет.'
    ws['M2'] = 'Больные'
    ws['N2'] = 'Командировка'
    ws['O2'] = 'Другие причины'

    row = 3
    for r in res_list:
        ws.cell(column=1, row=row, value=r['dep_unique_id']).border = br2
        ws.cell(column=2, row=row, value=r['num_fighter']).border = br2
        ws.cell(column=3, row=row, value=r['karaul_list']).border = br2
        ws.cell(column=4, row=row, value=r['karaul_all']).border = br2
        ws.cell(column=5, row=row, value=r['karaul_fighter']).border = br2
        ws.cell(column=6, row=row, value=r['head']).border = br2
        ws.cell(column=7, row=row, value=r['commander']).border = br2
        ws.cell(column=8, row=row, value=r['drivers']).border = br2
        ws.cell(column=9, row=row, value=r['fire_fighter']).border = br2
        ws.cell(column=10, row=row, value=r['call_assistant']).border = br2
        ws.cell(column=11, row=row, value=r['gdzs']).border = br2
        ws.cell(column=12, row=row, value=r['vacation']).border = br2
        ws.cell(column=13, row=row, value=r['sick']).border = br2
        ws.cell(column=14, row=row, value=r['business_trip']).border = br2
        ws.cell(column=15, row=row, value=r['others']).border = br2
        row += 1

    # SECOND SHEET

    ws1 = wb.create_sheet("Пожарная техника")

    for col_letter in alphabit[:7]:
        ws1.column_dimensions[col_letter] = ColumnDimension(ws1, bestFit=True, customWidth=True)
        ws1.column_dimensions[col_letter].width = 19
        ws1[col_letter + '1'].border = br
        ws1[col_letter + '2'].border = br
        ws1[col_letter + '1'].font = ft
        ws1[col_letter + '2'].font = ft
        ws1[col_letter + '1'].alignment = al
        ws1[col_letter + '2'].alignment = al

    ws1.row_dimensions[1].height = 20
    ws1.row_dimensions[2].height = 65

    # Merge Cells OF HEADER
    ws1.merge_cells('A1:A2')

    # Assignment
    ws1['A1'] = 'Наименование пожарных подразделений'
    ws1['B1'] = 'В расчете'
    ws1['B2'] = 'Тип основного пожарного автомобиля'
    ws1['C2'] = 'Марка специального пожарного автомобиля'
    ws1['D1'] = 'В резерве'
    ws1['D2'] = 'Тип основного пожарного автомобиля'
    ws1['E2'] = 'Марка специального пожарного автомобиля'
    ws1['F1'] = 'На ремонте'
    ws1['F2'] = 'Тип основного пожарного автомобиля'
    ws1['G2'] = 'Марка специального пожарного автомобиля'

    row = 3
    for r in res_list:
        tab_row = r['tab_row']
        if tab_row > -1:
            ws1.cell(column=1, row=row, value=r['dep_unique_id']).border = br2

            inner_row = row
            for i in range(tab_row + 1):
                try:
                    ws1.cell(column=2, row=inner_row, value=r['counting_brand'][i]).border = br2
                    ws1.cell(column=3, row=inner_row, value=r['counting_model'][i]).border = br2
                    inner_row += 1
                except IndexError:
                    ws1.cell(column=2, row=inner_row, value='').border = br2
                    ws1.cell(column=3, row=inner_row, value='').border = br2
                    inner_row += 1
            inner_row = row
            for i in range(tab_row + 1):
                try:
                    ws1.cell(column=4, row=inner_row, value=r['reserve_brand'][i]).border = br2
                    ws1.cell(column=5, row=inner_row, value=r['reserve_model'][i]).border = br2
                    inner_row += 1
                except IndexError:
                    ws1.cell(column=4, row=inner_row, value='').border = br2
                    ws1.cell(column=5, row=inner_row, value='').border = br2
                    inner_row += 1
            inner_row = row
            for i in range(tab_row + 1):
                try:
                    ws1.cell(column=6, row=inner_row, value=r['renovation_brand'][i]).border = br2
                    ws1.cell(column=7, row=inner_row, value=r['renovation_model'][i]).border = br2
                    inner_row += 1
                except IndexError:
                    ws1.cell(column=6, row=inner_row, value='').border = br2
                    ws1.cell(column=7, row=inner_row, value='').border = br2
                    inner_row += 1
            row = row + tab_row + 1
        else:
            ws1.cell(column=1, row=row, value=r['dep_unique_id']).border = br2
            for i in range(2, 8):
                ws1.cell(column=i, row=row, value='').border = br2

            row += 1

    wb.save(filename=full_path)
    return full_path